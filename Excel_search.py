# Po otrzymaniu ścieżki do katalogu, w którym znajdują się pliki Excel (xlsx), skrypt pyta o szukaną wartość
# po czym przeszukuje wszystkie pliki, a po odnalezieniu jej pierwszego wystąpienia zatrzymuje się
# i zwraca w odpowiedzi:
# ścieżkę, nazwę pliku, nazwę zakładki i numer komórki, w której szukana wartość występuje

import os
import openpyxl
import platform

def search_in_workbook(file_path, search_value):
    try:
        workbook = openpyxl.load_workbook(file_path, data_only=True)
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            for row in sheet.iter_rows():
                for cell in row:
                    if cell.value == search_value:
                        return file_path, sheet_name, cell.coordinate
    except Exception as e:
        print(f"Wystąpił błąd przy przetwarzaniu pliku {file_path}: {e}")
    return None

def search_in_directory(search_value, directory):
    found = False
    for file in os.listdir(directory):
        if file.endswith('.xlsx'):
            file_path = os.path.join(directory, file)
            result = search_in_workbook(file_path, search_value)
            if result:
                print(f"Znaleziono w pliku: {result[0]}, zakładka: {result[1]}, komórka: {result[2]}")
                found = True
                break  # Usuń tę linię, jeśli chcesz kontynuować szukanie w kolejnych plikach
    if not found:
        print("Nie znaleziono podanej wartości.")

if __name__ == "__main__":
    directory = input("Podaj ścieżkę do katalogu: ")
    search_value = input("Podaj szukaną wartość: ")
    search_in_directory(search_value, directory)

    # Oczekiwanie na naciśnięcie dowolnego klawisza przed zamknięciem
    if platform.system() == 'Windows':
        os.system('pause')  # Tylko dla Windows
    else:
        input("Press any button to continue...")  # Dla Unix/Linux i MacOS
