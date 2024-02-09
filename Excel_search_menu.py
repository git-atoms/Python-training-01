import os
import openpyxl
import platform

def search_in_workbook(file_path, search_value):
    results = []
    try:
        workbook = openpyxl.load_workbook(file_path, data_only=True)
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            for row in sheet.iter_rows():
                for cell in row:
                    if cell.value == search_value:
                        file_name = os.path.basename(file_path)
                        results.append((file_name, sheet_name, cell.coordinate))
    except Exception as e:
        print(f"Wystąpił błąd przy przetwarzaniu pliku {file_path}: {e}")
    return results

def search_in_directory(search_value, directory):
    found = False
    for file in os.listdir(directory):
        if file.endswith('.xlsx'):
            file_path = os.path.join(directory, file)
            results = search_in_workbook(file_path, search_value)
            for result in results:
                print(f"Znaleziono w pliku: {result[0]}, zakładka: {result[1]}, komórka: {result[2]}")
                found = True
    if not found:
        print("Nie znaleziono podanej wartości.")

def main_loop():
    directory = input("Podaj ścieżkę do katalogu: ")
    while True:
        search_value = input("Podaj szukaną wartość: ")
        search_in_directory(search_value, directory)

        print("Co Tomku robimy dalej?")
        print("1. Szukamy nadal tutaj")
        print("2. Szukamy w nowym katalogu")
        print("3. Kończymy szukanie")
        choice = input("Wybierz opcję (1, 2, 3): ")

        if choice == '1':
            continue
        elif choice == '2':
            directory = input("Podaj ścieżkę do nowego katalogu: ")
        elif choice == '3':
            break
        else:
            print("Nieprawidłowy wybór, kończymy szukanie.")
            break

if __name__ == "__main__":
    main_loop()

    # Oczekiwanie na naciśnięcie dowolnego klawisza przed zamknięciem dla systemów innych niż Windows
    if platform.system() != 'Windows':
        input("Press any button to continue...")  # Dla Unix/Linux i MacOS
