"""
Skrypt ma za zadanie znaleźć podaną mu wartość komórki z pliku Excel (xlsx).
W tym celu przeszukuje wszystkie pliki o rozszerzeniu .xlsx we wskazanym katalogu oraz schodzi
rekursywnie w dół do podkatalogów, które również przeszukuje i zwraca jako wynik:

1./ Nazwy wszystkich plików będących w każdym napotkanym katalogu, w których znajduje się poszukiwana przez nas wartość,
wraz ze ścieżką do każdego pliku.
2./ Nazwę zakładki w każdym pliku
3./ Numer komórki pierwszej odnalezionej (poszukiwanej) wartości w każdej zakładce
4./ Ilość kolejnych wystąpień w tej samej zakładce arkusza (jeśli występuje więcej niż jedna)
5./ Łączny czas wykonywania się skryptu

Na koniec zadaje pytanie o następną czynność, którą ma wykonać.
Jeśli odpowiedź będzie spoza zakresu 1-3 (takie są opcje do wyboru), zadaje pytanie ponownie, aż do otrzymania
właściwej odpowiedzi.

"""

import os
import openpyxl
import platform
import time

def search_in_workbook(file_path, search_value):
    results = {}
    try:
        workbook = openpyxl.load_workbook(file_path, data_only=True)
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            occurrences = 0
            for row in sheet.iter_rows():
                for cell in row:
                    if cell.value == search_value:
                        occurrences += 1
                        if sheet_name not in results:
                            results[sheet_name] = [(file_path, cell.coordinate)]
                        else:
                            results[sheet_name].append((file_path, cell.coordinate))
            if occurrences > 1:
                results[sheet_name] += (occurrences,)
    except Exception as e:
        print(f"Wystąpił błąd przy przetwarzaniu pliku {file_path}: {e}")
    return results

def search_in_directory(search_value, directory):
    start_time = time.time()
    found = False
    for root, dirs, files in os.walk(directory):
        for file in files:
            if file.endswith('.xlsx'):
                file_path = os.path.join(root, file)
                results = search_in_workbook(file_path, search_value)
                for sheet, infos in results.items():
                    for info in infos:
                        print(f"\nPełna ścieżka: {info[0]}\n  Nazwa zakładki: {sheet}\n  Numer komórki: {info[1]}")
                        found = True
    if not found:
        print("Nie znaleziono podanej wartości.")
    print(f"\n        Łączny czas wyszukiwania: {time.time() - start_time:.2f} sekund.")

def menu():
    while True:
        print("\n\nCo Tomku robimy dalej?")
        print("1. Szukamy tutaj")
        print("2. Szukamy w nowym katalogu")
        try:
            choice = int(input())
            if 1 <= choice <= 2:
                return choice
        except ValueError:
            pass

def main_loop():
    directory = input("Podaj ścieżkę do katalogu: ")
    while True:
        search_value = input("Podaj szukaną wartość: ")
        search_in_directory(search_value, directory)

        choice = menu()
        if choice == 1:
            continue
        elif choice == 2:
            directory = input("Podaj ścieżkę do nowego katalogu: ")

if __name__ == "__main__":
    main_loop()

    # if platform.system() != 'Windows':
    #     input("Press any button to continue...")  # Dla Unix/Linux i MacOS

    # Oczekiwanie na naciśnięcie dowolnego klawisza przed zamknięciem
    if platform.system() == 'Windows':
        os.system('pause')  # Tylko dla Windows
    else:
        input("Press any button to continue...")  # Dla Unix/Linux i MacOS