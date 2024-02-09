import os
import openpyxl
import platform
import time


def search_in_workbook(file_path, search_value):
    results = {}
    try:
        workbook = openpyxl.load_workbook(file_path, data_only=True)
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            occurrences = 0
            for row in sheet.iter_rows():
                for cell in row:
                    if cell.value == search_value:
                        occurrences += 1
                        if sheet_name not in results:
                            file_name = os.path.basename(file_path)
                            results[sheet_name] = (file_name, cell.coordinate)
            if occurrences > 1:
                results[sheet_name] += (occurrences,)
    except Exception as e:
        print(f"Wystąpił błąd przy przetwarzaniu pliku {file_path}: {e}")
    return results


def search_in_directory(search_value, directory):
    start_time = time.time()
    found = False
    for file in os.listdir(directory):
        if file.endswith('.xlsx'):
            file_path = os.path.join(directory, file)
            results = search_in_workbook(file_path, search_value)
            for sheet, info in results.items():
                print(f"\n{info[0]}\n  Nazwa zakładki: {sheet}\n  Numer komórki: {info[1]}")
                # print(f"Znaleziono w pliku: {info[0]}, zakładka: {sheet}, komórka: {info[1]}")
                if len(info) == 3:
                    print(f"  Liczba wystąpień w zakładce: {info[2]}")
                print()  # Dodatkowy pusty wiersz po każdym wyniku
                found = True
    if not found:
        print("Nie znaleziono podanej wartości.")
    print(f"        Łączny czas wyszukiwania: {time.time() - start_time:.2f} sekund.")


def menu():
    while True:
        print("\n\nCo Tomku robimy dalej?")
        print("1. Szukamy tutaj")
        print("2. Szukamy w nowym katalogu")
        try:
            choice = int(input())
            if 1 <= choice <= 2:
                return choice
        except ValueError:
            pass


def main_loop():
    directory = input("Podaj ścieżkę do katalogu: ")
    while True:
        search_value = input("Podaj szukaną wartość: ")
        search_in_directory(search_value, directory)

        choice = menu()
        if choice == 1:
            continue
        elif choice == 2:
            directory = input("Podaj ścieżkę do nowego katalogu: ")


if __name__ == "__main__":
    main_loop()

    # Oczekiwanie na naciśnięcie dowolnego klawisza przed zamknięciem dla systemów innych niż Windows
    if platform.system() != 'Windows':
        input("Press any button to continue...")  # Dla Unix/Linux i MacOS
