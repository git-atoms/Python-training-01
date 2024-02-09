"""
Skrypt ma za zadanie znaleźć podaną mu wartość komórki z pliku Excel (xlsx).
W tym celu przeszukuje wszystkie pliki o rozszerzeniu .xlsx w katalogu i zwraca jako wynik:

1./ Nazwy wszystkich plików będących w tym katalogu, w których znajduje się poszukiwana przez nas wartość
2./ Nazwę zakładki w każdym pliku
3./ Numer komórki pierwszej odnalezionej (poszukiwanej) wartości w każdej zakładce
4./ Ilość kolejnych wystąpień w tej samej zakładce arkusza (jeśli występuje więcej niż jedna)
5./ Łączny czas wykonywania się skryptu

Na koniec zadaje pytanie o następną czynność, którą ma wykonać.
Jeśli odpowiedź będzie spoza zakresu 1-3 (takie są opcje do wyboru), zadaje pytanie ponownie, aż do otrzymania
właściwej odpowiedzi.

"""

import os
import openpyxl
import platform
import time


def search_in_workbook(file_path, search_value):
    results = {}
    try:
        workbook = openpyxl.load_workbook(file_path, data_only=True)
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            occurrences = 0
            for row in sheet.iter_rows():
                for cell in row:
                    if cell.value == search_value:
                        occurrences += 1
                        if sheet_name not in results:
                            file_name = os.path.basename(file_path)
                            results[sheet_name] = (file_name, cell.coordinate)
            if occurrences > 1:
                results[sheet_name] += (occurrences,)
    except Exception as e:
        print(f"Wystąpił błąd przy przetwarzaniu pliku {file_path}: {e}")
    return results


def search_in_directory(search_value, directory):
    start_time = time.time()
    found = False
    for file in os.listdir(directory):
        if file.endswith('.xlsx'):
            file_path = os.path.join(directory, file)
            results = search_in_workbook(file_path, search_value)
            for sheet, info in results.items():
                print(f"\n{info[0]}\n  Nazwa zakładki: {sheet}\n  Numer komórki: {info[1]}")
                if len(info) == 3:
                    print(f"  Liczba wystąpień w zakładce: {info[2]}")
                print()  # Dodatkowy pusty wiersz po każdym wyniku
                found = True
    if not found:
        print("Nie znaleziono podanej wartości.")
    print(f"      Łączny czas wyszukiwania: {time.time() - start_time:.2f} sekund.")


def main_loop():
    directory = input("Podaj ścieżkę do katalogu: ")
    while True:
        search_value = input("Podaj szukaną wartość: ")
        search_in_directory(search_value, directory)

        print("\n\nCo Tomku robimy dalej?")
        print("1. Szukamy tutaj")
        print("2. Szukamy w nowym katalogu")
        # print("3. Kończymy szukanie")

        choice = input()
        if choice == '1':
            continue
        elif choice == '2':
            directory = input("Podaj ścieżkę do nowego katalogu: ")
        # elif choice == '3':
        #     break


if __name__ == "__main__":
    main_loop()

    # Oczekiwanie na naciśnięcie dowolnego klawisza przed zamknięciem dla systemów innych niż Windows
    if platform.system() != 'Windows':
        input("Press any button to continue...")  # Dla Unix/Linux i MacOS
